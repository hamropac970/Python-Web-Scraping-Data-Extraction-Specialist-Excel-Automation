import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import random
import logging
from dataclasses import dataclass
from typing import List, Dict, Optional
from concurrent.futures import ThreadPoolExecutor
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - [%(levelname)s] - %(message)s',
    handlers=[logging.FileHandler("ultimate_god_scraper.log", encoding='utf-8'), logging.StreamHandler()]
)


@dataclass
class ProductData:
    title: str
    price: float
    availability: str


class GodModeScraper:

    def __init__(self, base_url: str, max_workers: int = 5) -> None:
        self.base_url = base_url
        self.max_workers = max_workers
        self.products: List[ProductData] = []

        self.user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15"
        ]

        self.proxies_pool: List[Dict[str, str]] = [
            # {"http": "http://username:password@proxy_ip:port"}
        ]

    def _get_headers(self) -> Dict[str, str]:
        return {"User-Agent": random.choice(self.user_agents)}

    def _get_proxy(self) -> Optional[Dict[str, str]]:
        return random.choice(self.proxies_pool) if self.proxies_pool else None

    def scrape_single_page(self, page_num: int) -> None:
        url = f"{self.base_url}index.html" if page_num == 1 else f"{self.base_url}page-{page_num}.html"
        logging.info(f"Loading page (via Thread): {page_num}")

        try:
            # Open a session and send a request with a proxy
            with requests.Session() as session:
                response = session.get(
                    url,
                    headers=self._get_headers(),
                    proxies=self._get_proxy(),
                    timeout=12
                )
                if response.status_code == 200:
                    response.encoding = 'utf-8'
                    self.parse_html(response.text)
                else:
                    logging.warning(f"Failed to load page: {page_num} (Status: {response.status_code})")
        except Exception as e:
            logging.error(f"Unexpected error on page {page_num}: {e}")

    def parse_html(self, html: str) -> None:
        soup = BeautifulSoup(html, 'html.parser')
        books = soup.find_all('article', class_='product_pod')

        for book in books:
            try:
                title = book.h3.a['title']
                price_raw = book.find('p', class_='price_color').text
                price_clean = ''.join(c for c in price_raw if c.isdigit() or c == '.')
                availability = book.find('p', class_='instock availability').text.strip()

                self.products.append(ProductData(title=title, price=float(price_clean), availability=availability))
            except Exception as e:
                pass

    def start_fast_scraping(self, total_pages: int) -> None:
        logging.info(f"Fast multi-threaded scraping started ({self.max_workers} threads).")
        start_time = time.time()

        pages = list(range(1, total_pages + 1))
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            executor.map(self.scrape_single_page, pages)

        end_time = time.time()
        logging.info(f"All pages scraped in {end_time - start_time:.2f} seconds!")
        self.export_to_god_excel()

    def export_to_god_excel(self, filename: str = 'god_mode_report.xlsx') -> None:
        if not self.products:
            logging.warning("No data collected to export.")
            return

        df = pd.DataFrame([p.__dict__ for p in self.products])
        df.columns = ['Product Name', 'Price (£)', 'Availability']

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Main Data')

            workbook = writer.book
            worksheet = writer.sheets['Main Data']

            header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='112233', end_color='112233', fill_type='solid')  # Midnight Blue
            data_font = Font(name='Segoe UI', size=10)
            thin_border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
                                 top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))


            for col_num in range(1, 4):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')


            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='left')
                worksheet.cell(row=row, column=2).alignment = Alignment(horizontal='right')
                worksheet.cell(row=row, column=3).alignment = Alignment(horizontal='center')
                for col in range(1, 4):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = thin_border


            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                worksheet.column_dimensions[col[0].column_letter].width = max(max_len + 4, 15)


            # Build a new analytics panel on the right side, leaving an empty column
            worksheet.cell(row=2, column=5, value="MINI ANALYTICS").font = Font(name='Segoe UI', size=12, bold=True,
                                                                                color='112233')

            # Write AUTOMATIC FORMULAS directly into Excel
            worksheet.cell(row=4, column=5, value="Average Price:")
            worksheet.cell(row=4, column=6, value=f"=AVERAGE(B2:B{worksheet.max_row})").font = Font(bold=True)

            worksheet.cell(row=5, column=5, value="Highest Price:")
            worksheet.cell(row=5, column=6, value=f"=MAX(B2:B{worksheet.max_row})").font = Font(bold=True)

            worksheet.cell(row=6, column=5, value="Total Items:")
            worksheet.cell(row=6, column=6, value=f"=COUNTA(A2:A{worksheet.max_row})").font = Font(bold=True)

        logging.info(f"PERFECT FINISH! '{filename}' is ready with built-in analytics.")


if __name__ == "__main__":
    URL = "https://books.toscrape.com/catalogue/category/books/science_22/"

    # max_workers=5 means the program runs in parallel like 5 browsers at once
    scraper = GodModeScraper(base_url=URL, max_workers=5)
    scraper.start_fast_scraping(total_pages=1)